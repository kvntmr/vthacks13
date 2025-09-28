"""
Excel Parser Service
Extracts data from Excel files (.xlsx, .xls) with improved AI-readable formatting
"""

import os
import io
from typing import Dict, Any, List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

class ExcelParser:
    """Parser for Excel files with enhanced AI-readable formatting"""
    
    def __init__(self):
        """Initialize the Excel parser"""
        pass
    
    def _create_structured_text(self, data_list: List[List[str]], sheet_name: str, df: pd.DataFrame) -> str:
        """
        Create table-formatted text representation of Excel data for AI recognition
        
        Args:
            data_list: List of rows with cell data
            sheet_name: Name of the worksheet
            df: Pandas DataFrame for dimensions
            
        Returns:
            Table-formatted text representation
        """
        text_lines = []
        
        # Add worksheet header
        text_lines.append(f"=== WORKSHEET: {sheet_name} ===")
        text_lines.append(f"Dimensions: {len(df)} rows × {len(df.columns)} columns")
        text_lines.append("")
        
        # Process data as a table
        if len(data_list) > 0:
            # Try to identify header row (first non-empty row)
            header_row = None
            data_start_row = 0
            
            for i, row in enumerate(data_list):
                if any(cell.strip() for cell in row):
                    if header_row is None:
                        header_row = row
                        data_start_row = i + 1
                    else:
                        break
            
            # Create table format
            if header_row and any(cell.strip() for cell in header_row):
                # Clean header row
                headers = [cell.strip() for cell in header_row if cell.strip()]
                
                # Create table header
                header_line = " | ".join(headers)
                separator_line = "-" * len(header_line)
                text_lines.append(header_line)
                text_lines.append(separator_line)
                
                # Add data rows
                for i, row in enumerate(data_list[data_start_row:], start=data_start_row):
                    if any(cell.strip() for cell in row):
                        # Align row data with headers
                        row_data = []
                        for j, cell in enumerate(row):
                            if j < len(headers):
                                row_data.append(cell.strip() if cell.strip() else "")
                            else:
                                break
                        
                        # Pad row if it's shorter than headers
                        while len(row_data) < len(headers):
                            row_data.append("")
                        
                        row_line = " | ".join(row_data)
                        text_lines.append(row_line)
            else:
                # No clear header, format as simple table
                text_lines.append("DATA (No clear headers detected):")
                for i, row in enumerate(data_list):
                    if any(cell.strip() for cell in row):
                        row_line = " | ".join(cell.strip() for cell in row if cell.strip())
                        text_lines.append(row_line)
        
        return "\n".join(text_lines)
    
    async def parse_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse an Excel file and extract data
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing parsed content
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Excel file not found: {file_path}")
            
            result = {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "file_type": "excel",
                "worksheets": [],
                "extracted_text": "",
                "metadata": {},
                "processing_summary": {
                    "total_worksheets": 0,
                    "total_cells_with_data": 0,
                    "total_text_length": 0
                }
            }
            
            # Load workbook for metadata
            workbook = load_workbook(file_path, data_only=True)
            
            # Extract metadata
            result["metadata"] = {
                "creator": workbook.properties.creator or "",
                "title": workbook.properties.title or "",
                "subject": workbook.properties.subject or "",
                "description": workbook.properties.description or "",
                "keywords": workbook.properties.keywords or "",
                "last_modified_by": workbook.properties.lastModifiedBy or "",
                "created": str(workbook.properties.created) if workbook.properties.created else "",
                "modified": str(workbook.properties.modified) if workbook.properties.modified else "",
                "version": workbook.properties.version or ""
            }
            
            # Extract data from each worksheet
            all_text = []
            total_cells = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet_data = {
                    "sheet_name": sheet_name,
                    "data": [],
                    "text_content": "",
                    "dimensions": {
                        "max_row": 0,
                        "max_column": 0
                    }
                }
                
                # Read worksheet with pandas for easier data handling
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    
                    # Get dimensions
                    worksheet_data["dimensions"] = {
                        "max_row": len(df),
                        "max_column": len(df.columns)
                    }
                    
                    # Convert to list of lists for consistency
                    data_list = df.fillna("").astype(str).values.tolist()
                    worksheet_data["data"] = data_list
                    
                    # Create structured text representation
                    worksheet_data["text_content"] = self._create_structured_text(data_list, sheet_name, df)
                    all_text.append(worksheet_data['text_content'])
                    
                    # Count non-empty cells
                    non_empty_cells = sum(1 for row in data_list for cell in row if cell.strip())
                    total_cells += non_empty_cells
                    
                except Exception as e:
                    worksheet_data["error"] = f"Failed to read worksheet {sheet_name}: {str(e)}"
                
                result["worksheets"].append(worksheet_data)
            
            # Combine all text
            result["extracted_text"] = "\n\n".join(all_text)
            result["processing_summary"] = {
                "total_worksheets": len(workbook.sheetnames),
                "total_cells_with_data": total_cells,
                "total_text_length": len(result["extracted_text"])
            }
            
            return result
            
        except Exception as e:
            return {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "file_type": "excel",
                "error": f"Failed to parse Excel file: {str(e)}",
                "extracted_text": "",
                "worksheets": [],
                "metadata": {},
                "processing_summary": {
                    "total_worksheets": 0,
                    "total_cells_with_data": 0,
                    "total_text_length": 0
                }
            }
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported file formats"""
        return [".xlsx", ".xls"]
    
    def is_supported(self, file_path: str) -> bool:
        """Check if file format is supported"""
        file_ext = os.path.splitext(file_path.lower())[1]
        return file_ext in self.get_supported_formats()
    
    async def parse_file_from_bytes(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse an Excel file from bytes and extract data
        
        Args:
            file_content: Excel file content as bytes
            filename: Name of the file
            
        Returns:
            Dictionary containing parsed content
        """
        try:
            result = {
                "file_path": filename,
                "file_name": filename,
                "file_type": "excel",
                "worksheets": [],
                "extracted_text": "",
                "metadata": {},
                "processing_summary": {
                    "total_worksheets": 0,
                    "total_cells_with_data": 0,
                    "total_text_length": 0
                }
            }
            
            # Create BytesIO object for pandas and openpyxl
            file_buffer = io.BytesIO(file_content)
            
            # Load workbook for metadata
            workbook = load_workbook(file_buffer, data_only=True)
            
            # Extract metadata
            result["metadata"] = {
                "creator": workbook.properties.creator or "",
                "title": workbook.properties.title or "",
                "subject": workbook.properties.subject or "",
                "description": workbook.properties.description or "",
                "keywords": workbook.properties.keywords or "",
                "last_modified_by": workbook.properties.lastModifiedBy or "",
                "created": str(workbook.properties.created) if workbook.properties.created else "",
                "modified": str(workbook.properties.modified) if workbook.properties.modified else "",
                "version": workbook.properties.version or ""
            }
            
            # Extract data from each worksheet
            all_text = []
            total_cells = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet_data = {
                    "sheet_name": sheet_name,
                    "data": [],
                    "text_content": "",
                    "dimensions": {
                        "max_row": 0,
                        "max_column": 0
                    }
                }
                
                # Read worksheet with pandas for easier data handling
                try:
                    # Reset buffer position for pandas
                    file_buffer.seek(0)
                    df = pd.read_excel(file_buffer, sheet_name=sheet_name, header=None)
                    
                    # Get dimensions
                    worksheet_data["dimensions"] = {
                        "max_row": len(df),
                        "max_column": len(df.columns)
                    }
                    
                    # Convert to list of lists for consistency
                    data_list = df.fillna("").astype(str).values.tolist()
                    worksheet_data["data"] = data_list
                    
                    # Create structured text representation
                    worksheet_data["text_content"] = self._create_structured_text(data_list, sheet_name, df)
                    all_text.append(worksheet_data['text_content'])
                    
                    # Count non-empty cells
                    non_empty_cells = sum(1 for row in data_list for cell in row if cell.strip())
                    total_cells += non_empty_cells
                    
                except Exception as e:
                    worksheet_data["error"] = f"Failed to read worksheet {sheet_name}: {str(e)}"
                
                result["worksheets"].append(worksheet_data)
            
            # Combine all text
            result["extracted_text"] = "\n\n".join(all_text)
            result["processing_summary"] = {
                "total_worksheets": len(workbook.sheetnames),
                "total_cells_with_data": total_cells,
                "total_text_length": len(result["extracted_text"])
            }
            
            return result
            
        except Exception as e:
            return {
                "file_path": filename,
                "file_name": filename,
                "file_type": "excel",
                "error": f"Failed to parse Excel file: {str(e)}",
                "extracted_text": "",
                "worksheets": [],
                "metadata": {},
                "processing_summary": {
                    "total_worksheets": 0,
                    "total_cells_with_data": 0,
                    "total_text_length": 0
                }
            }